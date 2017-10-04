import os                                          ##importar libreria de comandos del sistema windows CMD
from tqdm import tqdm                              ##libreria para mostrar barra de progreso
from openpyxl import load_workbook                 ##Se el modulo de lectura de archivos (load_workbook ) de la libreria openpyxl
     
MotionModel = []                                   #Modelo de asignacion vacio para seleccion de modelo de movmiento de placa

##Valores indicativos de las celdas de los modelos de movmiento de placas tectonicas
GSMR2_1        = ['B4','C4','D3','E3',           "GSMR2_1"]
NNR_MORVEL    = ['B6','C6','D5','E5',         "NNR_MORVEL"]
GEODVEL2010    = ['B8','C8','D7','E7',       "GEODVEL2010"]
MORVEL2010     = ['B10','C10','D9','E9',      "MORVEL2010"]
ITRF2008       = ['B12','C12','D11','E11',      "ITRF2008"]
APKIM2005_DGFI = ['B14','C14','D13','E13',"APKIM2005_DGFI"]
APKIM2005_IGN  = ['B16','C16','D15','E15', "APKIM2005_IGN"]
GSMR1_2       = ['B18','C18','D17','E17',        "GSMR1_2"]
CGPS2004       = ['B20','C20','D19','E19',      "CGPS2004"]
REVEL2000      = ['B22','C22','D21','E21',     "REVEL2000"]
ITRF2000AS     = ['B24','C24','D23','E23',    "ITRF2000AS"]
HS3_NUVEL1A    = ['B26','C26','D25','E25',   "HS3_NUVEL1A"]
APKIM2000      = ['B28','C28','D27','E27',     "APKIM2000"]
ITRF2000DA     = ['B30','C30','D29','E29',    "ITRF2000DA"]
HS2_NUVEL1A    = ['B32','C32','D31','E31',   "HS2_NUVEL1A"]
NUVEL1A        = ['B34','C34','D33','E33',       "NUVEL1A"]
NUVEL1         = ['B36','C36','D35','E35',        "NUVEL1"]

os.chdir("D:/ESTACIONES NZ(SA)/SIRGAS")                     ##Asignacion de la ruta donde se encuentran los archivos xlsx
ruta = os.getcwd()                                            #Obtiene de ubicacion de ruta actual en el programa
print ("DIRECCION: " + ruta)                                  ##imprime la direccion actual del programa mediante el metodo getcwd()
       
                                    ###FRAGMENTO IMPORTANTE, EN CASO DE ERROR COMPRUEBE QUE LAS DIRECCIONES ESTEN BIEN ESCRITAS###
#///////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////#
MotionModel = NUVEL1                                      ##Asignacion de los datos de modelo especifico a la variable de uso temporal                                              #
PathDir = 'C:/Users/julia/Desktop/mapas/Mapas con vectores de velocidad - Colombia/MAPAS - NZ(SA)' #ruta especificada en la cual se creara el directorio que contendra el archivo .gmt #
#///////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////////#

#os.mkdir(PathDir + '/MAPA - ' + MotionModel[4])               #comando para crear directorio de guardado del archivo en una ruta especifica
rutaGMT  = PathDir + '/MAPA - ' + MotionModel[4] + "/"        #Asignacion de la ruta del directorio creado a una variable        
fichero = open(rutaGMT+"SIRGAS(" + MotionModel[4]+").gmt", "w")            #se crea un archivo fichero de extension .gmt en la direccion indicada por la variable
 
print("archivos encontrados: " + str(len(os.listdir("."))))

for file in tqdm(os.listdir(".")):                             ##conteo de archivos xlsx y muestreo de la barra de progreso
    if (file != "script_xls.py" and file != "script-xls.pyproj" ):            ##se exeptua el procesamiento de estos dos archivos ya que son el programa y no son archivos de formato xlsx

        wb = load_workbook(file)                               ##Comando para cargar archivos.xlsx
        sheetname = str(wb.get_sheet_names())                  ##Comando para obtener el nombre de las hojas de calculo del archivo y convertirlo en string
        sheetname = sheetname[2:-2]                            ##Se ajusta el nombre del sheetname ya que viene con estos caracteres de mas ['sheetname']
        name_station = sheetname[0:4]                          ##Al recortar el sheetname se puede obtener el nombre de la estacion.
        #print (name_station)   
        sheet_ranges = wb[sheetname]                           ##De esta forma se guarda la hoja de calculo dentro de una variable

##ahora mediante el nombre de la hoja de calculo, se asigna la celda de interes, -->>
##-->>luego al usar el .value se podra ver el valor de la celda que esta contiene.

        long = str(sheet_ranges[MotionModel[0]].value)
        long = long[:-1]
        latd = str(sheet_ranges[MotionModel[1]].value)
        latd = latd[:-1]
        Evel = sheet_ranges[MotionModel[2]].value                
        Nvel = sheet_ranges[MotionModel[3]].value                

##impresion y guardado en fichero los datos provenientes de las celdas de interes 
       # print(long + " " + str(latd) + " " + str(Evel) + " " + str(Nvel) + " " + name_station)              
        fichero.write(long + " " + str(latd) + " " + str(Evel) + " " + str(Nvel) + " " + name_station + str("\n"))
      
fichero.close()   
print("DATA(" + MotionModel[4]+").gmt" + " ha sido creado")                             